# -*- coding: utf-8 -*-
"""
1. Adiciona CRE_Cod e CRE_Nome em Matriculas por Municipio, Infra por Municipio, Docentes por Municipio
2. Cria nova aba 'Escolas (Lista 2025)' com: CO_ENTIDADE, NO_ENTIDADE, CO_MUNICIPIO, NO_MUNICIPIO, CRE_Cod, CRE_Nome, TP_DEPENDENCIA, TP_SITUACAO_FUNCIONAMENTO, LATITUDE, LONGITUDE
Salva como Dados_Municipios_RS_Painel_v2.xlsx (feche o original antes de substituir)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import pandas as pd
import json
import os
from openpyxl.styles import PatternFill, Font, Alignment

BASE = r"C:\Users\mathe\OneDrive\Desktop\Trabalhos\02. Joinville\25. Painel de Indicadores Abertos Joinville\04. Produto 4_Indicadores Educacionais"
EXCEL_PATH  = os.path.join(BASE, "Dados_Municipios_RS_Painel.xlsx")
OUT_PATH    = os.path.join(BASE, "Dados_Municipios_RS_Painel_v2.xlsx")
LOOKUP_PATH = os.path.join(BASE, "painel", "dados", "rs_cre_lookup.json")
ESCOLA_CSV  = os.path.join(BASE,
    r"00. Bases de Dados\01. Acesso e Matrículas (Fonte Censo Escolar_Inep_2019_2025)"
    r"\01. extrações_2019_2025\Tabela_Escola_2025.csv")

# ── CRE lookup ──
with open(LOOKUP_PATH, 'r', encoding='utf-8') as f:
    lookup = json.load(f)
mun_to_cre = {str(k).zfill(7): v for k, v in lookup['mun_to_cre'].items()}

def get_cre(ibge, field):
    return mun_to_cre.get(str(ibge).split('.')[0].zfill(7), {}).get(field, '')

# ── Load original workbook ──
xl = pd.ExcelFile(EXCEL_PATH)
all_dfs = {s: pd.read_excel(EXCEL_PATH, sheet_name=s) for s in xl.sheet_names}

# ── Add CRE to municipality sheets ──
MUN_SHEETS = {
    'Matriculas por Municipio': 'Código IBGE',
    'Infra por Municipio':      'Codigo IBGE',
    'Docentes por Municipio':   'Codigo IBGE',
}
for sheet, ibge_col in MUN_SHEETS.items():
    df = all_dfs[sheet].copy()
    # Remove old CRE cols if they exist
    df = df[[c for c in df.columns if c not in ('CRE_Cod', 'CRE_Nome')]]
    pos = list(df.columns).index(ibge_col)
    df.insert(pos + 1, 'CRE_Cod',  df[ibge_col].map(lambda x: get_cre(x, 'cod_cre')))
    df.insert(pos + 2, 'CRE_Nome', df[ibge_col].map(lambda x: get_cre(x, 'nome_cre')))
    print(f"  '{sheet}': {(df['CRE_Cod'] != '').sum()}/{len(df)} mapeadas")
    all_dfs[sheet] = df

# ── Build school list tab ──
print(f"\nCarregando escolas de 2025...")
df_escola = pd.read_csv(ESCOLA_CSV, sep=';', encoding='latin-1',
    usecols=['NU_ANO_CENSO','CO_ENTIDADE','NO_ENTIDADE','CO_MUNICIPIO','NO_MUNICIPIO',
             'TP_DEPENDENCIA','TP_SITUACAO_FUNCIONAMENTO','LATITUDE','LONGITUDE',
             'TP_LOCALIZACAO'])

# Keep only Rio Grande do Sul (SG_UF = RS, CO_UF = 43)
# Re-read with SG_UF to filter
df_escola2 = pd.read_csv(ESCOLA_CSV, sep=';', encoding='latin-1',
    usecols=['NU_ANO_CENSO','SG_UF','CO_ENTIDADE','NO_ENTIDADE','CO_MUNICIPIO','NO_MUNICIPIO',
             'TP_DEPENDENCIA','TP_SITUACAO_FUNCIONAMENTO','LATITUDE','LONGITUDE','TP_LOCALIZACAO'])
df_rs = df_escola2[df_escola2['SG_UF'] == 'RS'].copy()
print(f"  Escolas RS 2025: {len(df_rs)}")

# Add CRE
df_rs['CRE_Cod']  = df_rs['CO_MUNICIPIO'].map(lambda x: get_cre(x, 'cod_cre'))
df_rs['CRE_Nome'] = df_rs['CO_MUNICIPIO'].map(lambda x: get_cre(x, 'nome_cre'))

# Rename for clarity
df_rs = df_rs.rename(columns={
    'NU_ANO_CENSO': 'Ano',
    'CO_ENTIDADE':  'Cod_Escola',
    'NO_ENTIDADE':  'Nome_Escola',
    'CO_MUNICIPIO': 'Codigo_IBGE',
    'NO_MUNICIPIO': 'Municipio',
    'TP_DEPENDENCIA': 'Dependencia',   # 1=Fed,2=Est,3=Mun,4=Priv
    'TP_SITUACAO_FUNCIONAMENTO': 'Situacao',  # 1=Ativa
    'TP_LOCALIZACAO': 'Localizacao',   # 1=Urbana, 2=Rural
})
# Keep ativas only (Situacao == 1)
df_rs = df_rs[df_rs['Situacao'] == 1].copy()
# Keep relevant columns in order
df_rs = df_rs[['Ano','Cod_Escola','Nome_Escola','Codigo_IBGE','Municipio',
                'CRE_Cod','CRE_Nome','Dependencia','Localizacao','LATITUDE','LONGITUDE']]
df_rs = df_rs.sort_values(['CRE_Cod','Municipio','Nome_Escola']).reset_index(drop=True)
print(f"  Escolas ativas RS: {len(df_rs)}")
print(f"  CREs mapeadas: {(df_rs['CRE_Cod'] != '').sum()}/{len(df_rs)}")

all_dfs['Escolas (Lista 2025)'] = df_rs

# ── Save ──
print(f"\nSalvando em: {OUT_PATH}")
with pd.ExcelWriter(OUT_PATH, engine='openpyxl') as writer:
    for sheet, df in all_dfs.items():
        df.to_excel(writer, sheet_name=sheet, index=False)
        ws = writer.sheets[sheet]
        # Style CRE and school name columns
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in ('CRE_Cod', 'CRE_Nome', 'Nome_Escola'):
                cell = ws.cell(row=1, column=col_idx)
                color = "005A32" if 'CRE' in col_name else "1A237E"
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')

print(f"✅ Salvo! {os.path.getsize(OUT_PATH)//1024} KB")
print(f"   Renomeie para Dados_Municipios_RS_Painel.xlsx quando fechar o original.")
