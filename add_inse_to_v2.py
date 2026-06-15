# -*- coding: utf-8 -*-
"""
Adiciona abas INSE na planilha Dados_Municipios_RS_Painel_v2.xlsx
Novas abas:
  - INSE (Serie Temporal): estatísticas estaduais 2019 e 2023
  - INSE por Municipio: INSE médio, nível, distribuição por município
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import pandas as pd
import numpy as np
import json, os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers

BASE = r"C:\Users\mathe\OneDrive\Desktop\Trabalhos\02. Joinville\25. Painel de Indicadores Abertos Joinville\04. Produto 4_Indicadores Educacionais"
XLSX_PATH = os.path.join(BASE, "Dados_Municipios_RS_Painel_v2.xlsx")
INSE_DIR  = os.path.join(BASE, "00. Bases de Dados", "04. Desigualdades Educacionais (INSE)")
LOOKUP_PATH = os.path.join(BASE, "painel", "dados", "rs_cre_lookup.json")

UF_COD = 43
DEP_ESTADUAL = 2

# CRE Lookup
with open(LOOKUP_PATH, 'r', encoding='utf-8') as f:
    lookup = json.load(f)
mun_to_cre = {str(k): v for k, v in lookup['mun_to_cre'].items()}

def get_cre(ibge, field):
    return mun_to_cre.get(str(int(ibge)), {}).get(field, '')

files = [
    ("2019", "INSE_2019_ESCOLAS.xlsx", "INSE_2019"),
    ("2023", "INSE_2023_escolas.xlsx", "INSE_ESC_2023"),
]

# ── Load and normalize INSE data ──
all_dfs = {}
for ano, fname, sheet in files:
    fpath = os.path.join(INSE_DIR, fname)
    print(f"Lendo {fname}...", end=" ", flush=True)
    df = pd.read_excel(fpath, sheet_name=sheet)
    
    renames = {
        'ID_ESCOLA': 'CO_ESCOLA', 'NO_ESCOLA': 'NOME_ESCOLA', 'NOME_ESCOLA': 'NOME_ESCOLA',
        'NO_MUNICIPIO': 'NOME_MUNICIPIO', 'NOME_MUNICIPIO': 'NOME_MUNICIPIO',
        'INSE_VALOR_ABSOLUTO': 'MEDIA_INSE', 'TP_TIPO_REDE': 'TP_DEPENDENCIA',
    }
    df.rename(columns={k:v for k,v in renames.items() if k in df.columns}, inplace=True)
    df = df[(df['CO_UF'] == UF_COD) & (df['TP_DEPENDENCIA'] == DEP_ESTADUAL)].copy()
    print(f"{len(df)} escolas")
    all_dfs[ano] = df

# ══════════════════════════════════════════════════════════
# ABA 1: INSE (Serie Temporal)
# ══════════════════════════════════════════════════════════
niveis_labels = ['Nível I','Nível II','Nível III','Nível IV','Nível V','Nível VI','Nível VII','Nível VIII']

rows_st = []
for ano, df in all_dfs.items():
    urb = df[df['TP_LOCALIZACAO'] == 1]
    rur = df[df['TP_LOCALIZACAO'] == 2]
    
    classif = df['INSE_CLASSIFICACAO'].value_counts()
    nivel_counts = {n: int(classif.get(n, 0)) for n in niveis_labels}
    
    row = {
        'Ano': int(ano),
        'Escolas': len(df),
        'Alunos Respondentes': int(df['QTD_ALUNOS_INSE'].fillna(0).sum()),
        'INSE Médio': round(df['MEDIA_INSE'].mean(), 2),
        'INSE Mediana': round(df['MEDIA_INSE'].median(), 2),
        'Desvio Padrão': round(df['MEDIA_INSE'].std(), 2),
        'INSE Mínimo': round(df['MEDIA_INSE'].min(), 2),
        'INSE Máximo': round(df['MEDIA_INSE'].max(), 2),
        'Urbana - Escolas': len(urb),
        'Urbana - INSE Médio': round(urb['MEDIA_INSE'].mean(), 2),
        'Rural - Escolas': len(rur),
        'Rural - INSE Médio': round(rur['MEDIA_INSE'].mean(), 2),
        'Gap Urbana-Rural': round(urb['MEDIA_INSE'].mean() - rur['MEDIA_INSE'].mean(), 2),
    }
    for n in niveis_labels:
        short = n.replace('Nível ', '')
        row[f'Nível {short} (n)'] = nivel_counts[n]
        row[f'Nível {short} (%)'] = round(nivel_counts[n] / len(df) * 100, 1)
    
    rows_st.append(row)

df_st = pd.DataFrame(rows_st)
print(f"\n✅ Aba 'INSE (Serie Temporal)': {len(df_st)} linhas x {len(df_st.columns)} colunas")

# ══════════════════════════════════════════════════════════
# ABA 2: INSE por Municipio (2019 e 2023 side by side)
# ══════════════════════════════════════════════════════════
# Collect all municipality codes
all_munis = set()
for ano, df in all_dfs.items():
    for cod in df['CO_MUNICIPIO'].unique():
        all_munis.add(int(cod))

rows_mun = []
for cod in sorted(all_munis):
    nome = None
    row = {
        'Codigo IBGE': cod,
        'CRE_Cod': get_cre(cod, 'cod_cre'),
        'CRE_Nome': get_cre(cod, 'nome_cre'),
    }
    
    for ano, df in all_dfs.items():
        df_m = df[df['CO_MUNICIPIO'] == cod]
        if len(df_m) > 0:
            if nome is None:
                nome = df_m['NOME_MUNICIPIO'].iloc[0] if 'NOME_MUNICIPIO' in df_m.columns else f'Cod.{cod}'
            row[f'INSE {ano}'] = round(df_m['MEDIA_INSE'].mean(), 2)
            row[f'Nível {ano}'] = df_m['INSE_CLASSIFICACAO'].mode().iloc[0]
            row[f'Escolas {ano}'] = len(df_m)
            row[f'Alunos {ano}'] = int(df_m['QTD_ALUNOS_INSE'].fillna(0).sum())
            
            # Level distribution
            classif = df_m['INSE_CLASSIFICACAO'].value_counts()
            for n in niveis_labels:
                short = n.replace('Nível ', '')
                row[f'{short} {ano} (n)'] = int(classif.get(n, 0))
        else:
            row[f'INSE {ano}'] = None
            row[f'Nível {ano}'] = None
            row[f'Escolas {ano}'] = 0
            row[f'Alunos {ano}'] = 0
            for n in niveis_labels:
                short = n.replace('Nível ', '')
                row[f'{short} {ano} (n)'] = 0
    
    row['Município'] = nome
    
    # Delta
    if row.get('INSE 2019') and row.get('INSE 2023'):
        row['Δ INSE 2019→2023'] = round(row['INSE 2023'] - row['INSE 2019'], 2)
    else:
        row['Δ INSE 2019→2023'] = None
    
    rows_mun.append(row)

df_mun = pd.DataFrame(rows_mun)

# Reorder columns for readability
cols_order = ['Codigo IBGE', 'CRE_Cod', 'CRE_Nome', 'Município',
              'INSE 2019', 'Nível 2019', 'Escolas 2019', 'Alunos 2019',
              'INSE 2023', 'Nível 2023', 'Escolas 2023', 'Alunos 2023',
              'Δ INSE 2019→2023']
# Add level distribution columns
for ano in ['2019', '2023']:
    for n in niveis_labels:
        short = n.replace('Nível ', '')
        cols_order.append(f'{short} {ano} (n)')

df_mun = df_mun[[c for c in cols_order if c in df_mun.columns]]
df_mun = df_mun.sort_values('INSE 2023', ascending=False, na_position='last').reset_index(drop=True)

print(f"✅ Aba 'INSE por Municipio': {len(df_mun)} linhas x {len(df_mun.columns)} colunas")

# ══════════════════════════════════════════════════════════
# WRITE TO EXISTING XLSX
# ══════════════════════════════════════════════════════════
print(f"\nAdicionando abas em: {XLSX_PATH}")

# Load existing workbook
wb = load_workbook(XLSX_PATH)

# Remove old INSE sheets if they exist
for name in ['INSE (Serie Temporal)', 'INSE por Municipio']:
    if name in wb.sheetnames:
        del wb[name]
        print(f"  Removida aba antiga: '{name}'")

# Style constants
HEADER_FILL_GREEN = PatternFill(start_color="005A32", end_color="005A32", fill_type="solid")
HEADER_FILL_BLUE  = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal='center', wrap_text=True)

def write_sheet(wb, name, df, header_fills=None):
    """Write a DataFrame as a new sheet with styled headers."""
    ws = wb.create_sheet(title=name)
    
    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        fill = HEADER_FILL_BLUE
        if header_fills:
            for key, f in header_fills.items():
                if key in col_name:
                    fill = f
                    break
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    
    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                cell.value = None
            else:
                cell.value = val
    
    # Auto-width (approx)
    for col_idx, col_name in enumerate(df.columns, 1):
        width = max(len(str(col_name)), 10)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(width + 2, 25)
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    print(f"  ✅ '{name}': {len(df)} linhas")

# Write sheets
write_sheet(wb, 'INSE (Serie Temporal)', df_st, {
    'CRE': HEADER_FILL_GREEN, 'Nível': HEADER_FILL_ORANGE, 'Gap': HEADER_FILL_ORANGE
})
write_sheet(wb, 'INSE por Municipio', df_mun, {
    'CRE': HEADER_FILL_GREEN, 'Δ': HEADER_FILL_ORANGE
})

# Save
wb.save(XLSX_PATH)
print(f"\n✅ Planilha salva! {os.path.getsize(XLSX_PATH)//1024} KB")
print(f"   Abas: {wb.sheetnames}")
