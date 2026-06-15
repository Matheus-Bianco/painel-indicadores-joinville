# -*- coding: utf-8 -*-
"""
ETL — Adequação da Formação Docente (AFD)
Processa planilhas INEP (2013-2025) e gera JSONs para o painel.
"""
import json, os, glob, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl

# ── Paths ──
BASE = r'c:\Users\mathe\OneDrive\Desktop\Trabalhos\02. Joinville\25. Painel de Indicadores Abertos Joinville\04. Produto 4_Indicadores Educacionais'
SRC  = os.path.join(BASE, '00. Bases de Dados', '09. Adequação da Formação Docente')
DST  = os.path.join(BASE, 'painel', 'dados')

# ── Column mapping ──
# 7 etapas × 5 grupos = 35 data columns (cols 9-43)
ETAPAS = [
    ('ed_inf',      9),   # Ed. Infantil: cols 9-13
    ('fund_total', 14),   # Fundamental Total: cols 14-18
    ('fund_ai',    19),   # Fund. Anos Iniciais: cols 19-23
    ('fund_af',    24),   # Fund. Anos Finais: cols 24-28
    ('medio',      29),   # Ensino Médio: cols 29-33
    ('eja_fund',   34),   # EJA Fundamental: cols 34-38
    ('eja_medio',  39),   # EJA Médio: cols 39-43
]

REDES = {
    'municipal': 'Municipal',
}

def safe_float(v):
    """Convert cell value to float, handling '--', None, etc."""
    if v is None or v == '--' or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def process_file(filepath):
    """Process a single AFD XLSX and return list of school dicts for RS only."""
    print(f"  Processando: {os.path.basename(filepath)}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    schools = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        # Filtra Joinville/SC (codigo IBGE 4209102)
        cod_mun = str(row[3])[:7] if row[3] else None
        if cod_mun != '4209102':
            continue
        
        school = {
            'ano': row[0],
            'cod_mun': str(row[3])[:7] if row[3] else None,
            'nome_mun': row[4],
            'cod_escola': str(row[5]) if row[5] else None,
            'nome_escola': row[6],
            'localizacao': row[7],  # Urbana/Rural
            'dependencia': row[8],  # Estadual/Municipal/etc
        }
        
        # Extract group percentages for each etapa
        for etapa_key, start_col in ETAPAS:
            groups = {}
            has_data = False
            for g in range(5):
                val = safe_float(row[start_col + g])
                groups[f'g{g+1}'] = val
                if val is not None:
                    has_data = True
            if has_data:
                school[etapa_key] = groups
        
        schools.append(school)
    
    wb.close()
    return schools

def aggregate_schools(schools, filter_dep=None):
    """
    Aggregate school-level data into state/municipality averages.
    Uses simple mean of percentages (standard INEP approach).
    """
    if filter_dep:
        schools = [s for s in schools if s['dependencia'] == filter_dep]
    
    # ── State-level aggregate ──
    state = {'total_escolas': len(schools)}
    for etapa_key, _ in ETAPAS:
        etapa_schools = [s for s in schools if etapa_key in s]
        if not etapa_schools:
            continue
        avg = {}
        for g in range(1, 6):
            vals = [s[etapa_key][f'g{g}'] for s in etapa_schools if s[etapa_key][f'g{g}'] is not None]
            avg[f'g{g}'] = round(sum(vals) / len(vals), 1) if vals else 0
        state[etapa_key] = avg
    
    # ── Municipality-level aggregate ──
    mun_groups = {}
    for s in schools:
        cod = s['cod_mun']
        if not cod:
            continue
        if cod not in mun_groups:
            mun_groups[cod] = []
        mun_groups[cod].append(s)
    
    mun_data = {}
    for cod, mun_schools in mun_groups.items():
        md = {'total_escolas': len(mun_schools)}
        for etapa_key, _ in ETAPAS:
            etapa_schools = [s for s in mun_schools if etapa_key in s]
            if not etapa_schools:
                continue
            avg = {}
            for g in range(1, 6):
                vals = [s[etapa_key][f'g{g}'] for s in etapa_schools if s[etapa_key][f'g{g}'] is not None]
                avg[f'g{g}'] = round(sum(vals) / len(vals), 1) if vals else 0
            md[etapa_key] = avg
        mun_data[cod] = md
    
    return state, mun_data

def build_lookup(all_schools):
    """Build municipality code → name lookup."""
    lookup = {}
    for s in all_schools:
        if s['cod_mun'] and s['nome_mun']:
            lookup[s['cod_mun']] = s['nome_mun']
    return lookup

def main():
    # Find all AFD files
    files = sorted(glob.glob(os.path.join(SRC, 'AFD_ESCOLAS_*.xlsx')))
    print(f"Encontrados {len(files)} arquivos AFD")
    
    # Process all files
    all_data = {}  # year → list of school dicts
    lookup_mun = {}
    
    for f in files:
        schools = process_file(f)
        if not schools:
            continue
        ano = str(schools[0]['ano'])
        all_data[ano] = schools
        lookup_mun.update(build_lookup(schools))
        print(f"    → {ano}: {len(schools)} escolas RS")
    
    anos = sorted(all_data.keys())
    print(f"\nAnos processados: {anos}")
    
    # ── Build main JSON (all networks) ──
    def build_json(filter_dep=None):
        result = {
            'serie_temporal': {},
            'por_municipio': {},
            'por_escola_2025': [],
            'lookup_municipios': lookup_mun,
        }
        
        for ano in anos:
            state, mun_data = aggregate_schools(all_data[ano], filter_dep)
            result['serie_temporal'][ano] = state
            result['por_municipio'][ano] = mun_data
        
        # Include school-level data for the last year
        ultimo = anos[-1]
        escolas_ultimo = all_data[ultimo]
        if filter_dep:
            escolas_ultimo = [s for s in escolas_ultimo if s['dependencia'] == filter_dep]
        result['por_escola_2025'] = escolas_ultimo
        
        return result
    
    # Main JSON (estadual network — default for the panel)
    print("\nGerando JSONs...")
    
    main_json = build_json('Municipal')
    with open(os.path.join(DST, '4_9_afd.json'), 'w', encoding='utf-8') as f:
        json.dump(main_json, f, ensure_ascii=False)
    print(f"  4_9_afd.json — {len(anos)} anos, {len(main_json['lookup_municipios'])} municípios")
    
    # Per-network JSONs
    for key, dep_name in REDES.items():
        net_json = build_json(dep_name)
        fname = f'4_9_afd_{key}.json'
        with open(os.path.join(DST, fname), 'w', encoding='utf-8') as f:
            json.dump(net_json, f, ensure_ascii=False)
        n_esc = net_json['serie_temporal'].get(anos[-1], {}).get('total_escolas', 0)
        print(f"  {fname} — {n_esc} escolas ({anos[-1]})")
    
    # All networks combined
    all_json = build_json(None)
    with open(os.path.join(DST, '4_9_afd_todas.json'), 'w', encoding='utf-8') as f:
        json.dump(all_json, f, ensure_ascii=False)
    n_esc = all_json['serie_temporal'].get(anos[-1], {}).get('total_escolas', 0)
    print(f"  4_9_afd_todas.json — {n_esc} escolas ({anos[-1]})")
    
    # ── Print summary for verification ──
    print("\n=== VERIFICAÇÃO ===")
    ultimo = anos[-1]
    st = main_json['serie_temporal'][ultimo]
    print(f"\nRede Municipal — {ultimo}:")
    print(f"  Total escolas: {st['total_escolas']}")
    for etapa_key, _ in ETAPAS:
        if etapa_key in st:
            e = st[etapa_key]
            print(f"  {etapa_key:12s}: G1={e['g1']:5.1f}%  G2={e['g2']:5.1f}%  G3={e['g3']:5.1f}%  G4={e['g4']:5.1f}%  G5={e['g5']:5.1f}%")
    
    print("\nDone!")

if __name__ == '__main__':
    main()
