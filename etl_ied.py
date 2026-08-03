# -*- coding: utf-8 -*-
"""ETL - Indicador de Esforco Docente (IED / INEP)."""
import sys
import io
import os
import json
import glob
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

BASE_PROJ = r"c:\Users\mathe\OneDrive\Desktop\Trabalhos\02. Joinville\25. Painel de Indicadores Abertos Joinville"

def find_src():
    for name in os.listdir(BASE_PROJ):
        if name.startswith("06.") and "Docente" in name:
            return os.path.join(BASE_PROJ, name)
    raise FileNotFoundError("Pasta 06.*Docente nao encontrada")

SRC = find_src()
OUT_DIR = os.path.join(BASE_PROJ, "04. Produto 4_Indicadores Educacionais", "painel", "dados")
CO_MUN = "4209102"

REDES = {
    "municipal": "Municipal",
    "estadual": "Estadual",
    "federal": "Federal",
    "particular": "Privada",
}

ETAPAS = [
    ("fund_total", 7),
    ("fund_ai", 13),
    ("fund_af", 19),
    ("medio", 25),
]

ELEVADO_RULE = {
    "fund_total": "n5_n6",
    "fund_ai": "n5_n6",
    "fund_af": "n6",
    "medio": "n6",
}


def safe_float(v):
    if v is None or v == "" or v == "--":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_niveis(row, start):
    niveis = {}
    has = False
    for i in range(6):
        val = safe_float(row[start + i] if start + i < len(row) else None)
        niveis["n%d" % (i + 1)] = val
        if val is not None:
            has = True
    return niveis if has else None


def add_elevado(etapa_key, niveis):
    if not niveis:
        return None
    out = dict(niveis)
    rule = ELEVADO_RULE.get(etapa_key, "n6")
    if rule == "n5_n6":
        a, b = out.get("n5"), out.get("n6")
        out["elevado"] = round((a or 0) + (b or 0), 1) if (a is not None or b is not None) else None
        out["elevado_regra"] = "Niveis 5 e 6"
    else:
        out["elevado"] = out.get("n6")
        out["elevado_regra"] = "Nivel 6"
    return out


def process_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        cod = str(row[3]).replace(".0", "") if row[3] is not None else ""
        if cod != CO_MUN:
            continue
        loc = str(row[5] or "").strip()
        dep = str(row[6] or "").strip()
        # normalize Publica (with or without accent)
        dep_ascii = dep.encode("ascii", "ignore").decode("ascii").lower()
        if dep_ascii == "publica":
            dep = "Publica"
        entry = {
            "ano": str(int(row[0])) if row[0] is not None else None,
            "localizacao": loc,
            "dependencia": dep,
            "etapas": {},
        }
        for key, start in ETAPAS:
            niv = parse_niveis(row, start)
            if niv:
                entry["etapas"][key] = add_elevado(key, niv)
        rows.append(entry)
    wb.close()
    return rows


def pick_row(rows, dependencia, localizacao="Total"):
    for r in rows:
        if r["dependencia"] == dependencia and r["localizacao"] == localizacao:
            return r
    return None


def build_rede(all_by_ano, dep_name):
    serie = {}
    por_loc = {}
    for ano in sorted(all_by_ano.keys()):
        rows = all_by_ano[ano]
        main = pick_row(rows, dep_name, "Total")
        if not main or not main["etapas"]:
            continue
        serie[ano] = dict(main["etapas"])
        loc_block = {}
        for loc in ("Urbana", "Rural"):
            lr = pick_row(rows, dep_name, loc)
            if lr and lr["etapas"]:
                loc_block[loc.lower()] = dict(lr["etapas"])
        if loc_block:
            por_loc[ano] = loc_block
    return serie, por_loc


def main():
    print("ETL Indicador de Esforco Docente (IED) - Joinville")
    print("  SRC:", SRC)
    files = sorted(glob.glob(os.path.join(SRC, "IED_MUNICIPIOS_*.xlsx")))
    print("  Arquivos:", len(files))
    all_by_ano = {}
    for fpath in files:
        print("  Lendo", os.path.basename(fpath), "...")
        rows = process_file(fpath)
        if not rows:
            print("    (sem Joinville)")
            continue
        ano = rows[0]["ano"]
        all_by_ano[ano] = rows
        deps = sorted({r["dependencia"] for r in rows})
        print("   ", ano, ":", len(rows), "linhas | deps=", deps)

    os.makedirs(OUT_DIR, exist_ok=True)
    anos = sorted(all_by_ano.keys())
    serie_total, _ = build_rede(all_by_ano, "Total")

    for key, dep_name in REDES.items():
        serie, por_loc = build_rede(all_by_ano, dep_name)
        payload = {
            "metadata": {
                "titulo": "Indicador de Esforco Docente (IED)",
                "fonte": "INEP - Indicador de Esforco Docente",
                "nota_tecnica": "Nota Tecnica n. 039/2014 e Nota CGCQTI/DEED/INEP n. 09/2016",
                "municipio": "Joinville",
                "cod_mun": CO_MUN,
                "rede": key,
                "dependencia_inep": dep_name,
                "anos": sorted(serie.keys()),
                "gerado_em": str(date.today()),
                "etapas": ["fund_total", "fund_ai", "fund_af", "medio"],
                "niveis": {
                    "1": "Ate 25 alunos; 1 turno, 1 escola, 1 etapa",
                    "2": "25 a 150 alunos; 1 turno, 1 escola, 1 etapa",
                    "3": "25 a 300 alunos; 1 ou 2 turnos; 1 escola e 1 etapa",
                    "4": "50 a 400 alunos; 2 turnos; 1-2 escolas; 2 etapas",
                    "5": "Mais de 300 alunos; 3 turnos; 2-3 escolas; 2-3 etapas",
                    "6": "Mais de 400 alunos; 3 turnos; 2-3 escolas; 2-3 etapas",
                },
                "regra_elevado": (
                    "Anos iniciais: % nos niveis 5 e 6. "
                    "Anos finais e Ensino Medio: % no nivel 6. "
                    "Fund. Total: soma dos niveis 5 e 6 (referencia)."
                ),
                "nota": (
                    "Percentual de funcoes docentes em cada nivel da escala IED. "
                    "Privada no Inep agrega particular e filantropica."
                    if key == "particular"
                    else "Percentual de funcoes docentes em cada nivel da escala IED (Censo Escolar)."
                ),
            },
            "serie_temporal": serie,
            "por_localizacao": por_loc,
            "serie_total_municipio": serie_total,
        }
        out = os.path.join(OUT_DIR, "4_13_ied_%s.json" % key)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        last = sorted(serie.keys())[-1] if serie else "-"
        elev_ai = serie.get(last, {}).get("fund_ai", {}).get("elevado")
        print("  OK", os.path.basename(out), "anos=", sorted(serie.keys()), "|", last, "AI elevado=", elev_ai)

    src = os.path.join(OUT_DIR, "4_13_ied_municipal.json")
    dst = os.path.join(OUT_DIR, "4_13_ied.json")
    if os.path.exists(src):
        with open(src, "r", encoding="utf-8") as f:
            data = json.load(f)
        with open(dst, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        print("  OK 4_13_ied.json (copia municipal)")

    print("Anos:", anos)
    print("Done.")


if __name__ == "__main__":
    main()
